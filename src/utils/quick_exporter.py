import pandas as pd
import streamlit as st
import io
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class QuickExporter:
    """Exportador rápido com templates otimizados e formatação avançada"""
    
    def __init__(self):
        self.cores = {
            'critico': 'FFEBEE',      # Vermelho claro
            'alto': 'FFF3E0',         # Laranja claro  
            'medio': 'F3E5F5',        # Roxo claro
            'baixo': 'E3F2FD',        # Azul claro
            'normal': 'F1F8E9',       # Verde claro
            'header': '2E86AB',       # Azul escuro
            'subheader': '90CAF9'     # Azul médio
        }
    
    def exportar_polo_excel(self, dados_polo: pd.DataFrame, nome_polo: str) -> bytes:
        """Exporta relatório completo do polo em Excel com formatação avançada"""
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Preparar dados principais
            colunas_principais = [
                'Ordem PagBank', 'Descricao_Urgencia', 'Dias_Em_Aberto', 
                'SLA Cliente', 'Status_SLA', 'Status da Ordem',
                'Último Tracking', 'Estado', 'Cidade', 'CEP'
            ]
            
            colunas_existentes = [col for col in colunas_principais if col in dados_polo.columns]
            dados_principais = dados_polo[colunas_existentes].copy()
            
            # Adicionar metadados
            dados_principais.insert(0, 'Data_Exportacao', datetime.now().strftime('%d/%m/%Y %H:%M'))
            dados_principais.insert(1, 'Polo_Responsavel', nome_polo)
            
            # ABA 1: Ordens Detalhadas
            dados_principais.to_excel(writer, sheet_name='Ordens_Detalhadas', index=False, startrow=2)
            
            # ABA 2: Resumo Executivo
            resumo = self._gerar_resumo_polo(dados_polo, nome_polo)
            resumo.to_excel(writer, sheet_name='Resumo_Executivo', index=False, startrow=2)
            
            # ABA 3: Apenas ordens críticas
            if 'Nivel_Urgencia' in dados_polo.columns:
                criticas = dados_polo[dados_polo['Nivel_Urgencia'] >= 4][colunas_existentes]
                if not criticas.empty:
                    criticas.insert(0, 'Data_Exportacao', datetime.now().strftime('%d/%m/%Y %H:%M'))
                    criticas.insert(1, 'Polo_Responsavel', nome_polo)
                    criticas.to_excel(writer, sheet_name='Ordens_Criticas', index=False, startrow=2)
            
            # ABA 4: Estatísticas Detalhadas
            stats = self._gerar_estatisticas_detalhadas(dados_polo)
            stats.to_excel(writer, sheet_name='Estatisticas', index=False, startrow=2)
            
            # ABA 5: Análise por Estado/Região
            if 'Estado' in dados_polo.columns:
                analise_geo = self._gerar_analise_geografica(dados_polo)
                analise_geo.to_excel(writer, sheet_name='Analise_Geografica', index=False, startrow=2)
            
            # Aplicar formatação avançada
            self._aplicar_formatacao_excel_avancada(writer, nome_polo, dados_polo)
        
        return output.getvalue()
    
    def exportar_polo_csv(self, dados_polo: pd.DataFrame, nome_polo: str) -> bytes:
        """Exporta lista simples do polo em CSV"""
        
        colunas_csv = [
            'Ordem PagBank', 'Dias_Em_Aberto', 'Status_SLA', 
            'Status da Ordem', 'Estado', 'Cidade'
        ]
        
        # Adicionar coluna de urgência se existir
        if 'Descricao_Urgencia' in dados_polo.columns:
            colunas_csv.insert(1, 'Descricao_Urgencia')
        
        colunas_existentes = [col for col in colunas_csv if col in dados_polo.columns]
        dados_csv = dados_polo[colunas_existentes].copy()
        
        # Adicionar metadados
        dados_csv.insert(0, 'Polo', nome_polo)
        dados_csv.insert(1, 'Data_Exportacao', datetime.now().strftime('%d/%m/%Y'))
        
        # Ordenar por urgência se disponível
        if 'Nivel_Urgencia' in dados_polo.columns:
            dados_csv = dados_csv.sort_values('Nivel_Urgencia', ascending=False)
        
        return dados_csv.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
    
    def exportar_resumo_executivo(self, dados_polo: pd.DataFrame, nome_polo: str) -> bytes:
        """Exporta apenas resumo executivo em formato CSV"""
        
        resumo_data = []
        
        # Métricas principais
        total_ordens = len(dados_polo)
        criticas = len(dados_polo[dados_polo['Nivel_Urgencia'] == 5]) if 'Nivel_Urgencia' in dados_polo.columns else 0
        altas = len(dados_polo[dados_polo['Nivel_Urgencia'] == 4]) if 'Nivel_Urgencia' in dados_polo.columns else 0
        medias = len(dados_polo[dados_polo['Nivel_Urgencia'] == 3]) if 'Nivel_Urgencia' in dados_polo.columns else 0
        vencidas = len(dados_polo[dados_polo['Status_SLA'] == 'Vencido']) if 'Status_SLA' in dados_polo.columns else 0
        atencao = len(dados_polo[dados_polo['Status_SLA'] == 'Atenção']) if 'Status_SLA' in dados_polo.columns else 0
        no_prazo = len(dados_polo[dados_polo['Status_SLA'] == 'No Prazo']) if 'Status_SLA' in dados_polo.columns else 0
        media_dias = dados_polo['Dias_Em_Aberto'].mean() if 'Dias_Em_Aberto' in dados_polo.columns else 0
        max_dias = dados_polo['Dias_Em_Aberto'].max() if 'Dias_Em_Aberto' in dados_polo.columns else 0
        
        # Seção: Informações Gerais
        resumo_data.extend([
            {'Polo': nome_polo, 'Categoria': 'INFORMAÇÕES GERAIS', 'Métrica': '', 'Valor': ''},
            {'Polo': nome_polo, 'Categoria': 'Identificação', 'Métrica': 'Nome do Polo', 'Valor': nome_polo},
            {'Polo': nome_polo, 'Categoria': 'Processamento', 'Métrica': 'Data/Hora da Exportação', 'Valor': datetime.now().strftime('%d/%m/%Y %H:%M')},
            {'Polo': nome_polo, 'Categoria': 'Volume', 'Métrica': 'Total de Ordens em Aberto', 'Valor': total_ordens},
            {'Polo': nome_polo, 'Categoria': '', 'Métrica': '', 'Valor': ''},
        ])
        
        # Seção: Distribuição por Urgência
        resumo_data.extend([
            {'Polo': nome_polo, 'Categoria': 'DISTRIBUIÇÃO POR URGÊNCIA', 'Métrica': '', 'Valor': ''},
            {'Polo': nome_polo, 'Categoria': 'Crítica', 'Métrica': '🔴 Ordens Críticas', 'Valor': criticas},
            {'Polo': nome_polo, 'Categoria': 'Alta', 'Métrica': '🟠 Ordens Alta Prioridade', 'Valor': altas},
            {'Polo': nome_polo, 'Categoria': 'Média', 'Métrica': '🟡 Ordens Média Prioridade', 'Valor': medias},
            {'Polo': nome_polo, 'Categoria': 'Percentual Crítico', 'Métrica': '% Ordens Críticas', 'Valor': f"{(criticas/total_ordens*100):.1f}%" if total_ordens > 0 else "0%"},
            {'Polo': nome_polo, 'Categoria': '', 'Métrica': '', 'Valor': ''},
        ])
        
        # Seção: Status SLA
        resumo_data.extend([
            {'Polo': nome_polo, 'Categoria': 'STATUS SLA', 'Métrica': '', 'Valor': ''},
            {'Polo': nome_polo, 'Categoria': 'Vencido', 'Métrica': 'Ordens com SLA Vencido', 'Valor': vencidas},
            {'Polo': nome_polo, 'Categoria': 'Atenção', 'Métrica': 'Ordens em Atenção', 'Valor': atencao},
            {'Polo': nome_polo, 'Categoria': 'No Prazo', 'Métrica': 'Ordens No Prazo', 'Valor': no_prazo},
            {'Polo': nome_polo, 'Categoria': 'Performance', 'Métrica': '% SLA Cumprido', 'Valor': f"{(no_prazo/total_ordens*100):.1f}%" if total_ordens > 0 else "0%"},
            {'Polo': nome_polo, 'Categoria': '', 'Métrica': '', 'Valor': ''},
        ])
        
        # Seção: Estatísticas de Tempo
        resumo_data.extend([
            {'Polo': nome_polo, 'Categoria': 'ESTATÍSTICAS DE TEMPO', 'Métrica': '', 'Valor': ''},
            {'Polo': nome_polo, 'Categoria': 'Média', 'Métrica': 'Média de Dias em Aberto', 'Valor': f"{media_dias:.1f}"},
            {'Polo': nome_polo, 'Categoria': 'Máximo', 'Métrica': 'Maior Tempo em Aberto', 'Valor': f"{max_dias} dias"},
            {'Polo': nome_polo, 'Categoria': '', 'Métrica': '', 'Valor': ''},
        ])
        
        # Seção: Top 15 ordens mais críticas
        if not dados_polo.empty and 'Dias_Em_Aberto' in dados_polo.columns:
            resumo_data.extend([
                {'Polo': nome_polo, 'Categoria': 'TOP 15 ORDENS MAIS CRÍTICAS', 'Métrica': '', 'Valor': ''},
            ])
            
            top_criticas = dados_polo.nlargest(15, 'Dias_Em_Aberto')
            for idx, (_, row) in enumerate(top_criticas.iterrows(), 1):
                status_sla = row.get('Status_SLA', 'N/A')
                estado = row.get('Estado', 'N/A')
                cidade = row.get('Cidade', 'N/A')
                
                resumo_data.append({
                    'Polo': nome_polo,
                    'Categoria': f'Top {idx:02d}',
                    'Métrica': f"Ordem {row['Ordem PagBank']} - {estado}/{cidade}", 
                    'Valor': f"{row['Dias_Em_Aberto']} dias - {status_sla}"
                })
        
        # Seção: Análise Geográfica (se disponível)
        if 'Estado' in dados_polo.columns:
            resumo_data.extend([
                {'Polo': nome_polo, 'Categoria': '', 'Métrica': '', 'Valor': ''},
                {'Polo': nome_polo, 'Categoria': 'DISTRIBUIÇÃO GEOGRÁFICA', 'Métrica': '', 'Valor': ''},
            ])
            
            dist_estados = dados_polo['Estado'].value_counts().head(10)
            for estado, qtd in dist_estados.items():
                resumo_data.append({
                    'Polo': nome_polo,
                    'Categoria': 'Geografia',
                    'Métrica': f'Estado {estado}',
                    'Valor': f"{qtd} ordens ({qtd/total_ordens*100:.1f}%)"
                })
        
        df_resumo = pd.DataFrame(resumo_data)
        return df_resumo.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
    
    def exportar_consolidado_todos_polos(self, relatorios_polo: Dict[str, pd.DataFrame]) -> bytes:
        """Exporta relatório consolidado de todos os polos"""
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # ABA 1: Resumo Geral
            resumo_geral = self._gerar_resumo_geral_polos(relatorios_polo)
            resumo_geral.to_excel(writer, sheet_name='Resumo_Geral', index=False, startrow=2)
            
            # ABA 2: Consolidado de Críticas
            todas_criticas = pd.DataFrame()
            for polo_id, dados_polo in relatorios_polo.items():
                if 'Nivel_Urgencia' in dados_polo.columns:
                    criticas_polo = dados_polo[dados_polo['Nivel_Urgencia'] >= 4].copy()
                    if not criticas_polo.empty:
                        criticas_polo['Polo_Origem'] = polo_id
                        todas_criticas = pd.concat([todas_criticas, criticas_polo], ignore_index=True)
            
            if not todas_criticas.empty:
                colunas_criticas = ['Polo_Origem', 'Ordem PagBank', 'Descricao_Urgencia', 'Dias_Em_Aberto', 'Status_SLA', 'Estado']
                colunas_existentes = [col for col in colunas_criticas if col in todas_criticas.columns]
                todas_criticas[colunas_existentes].to_excel(writer, sheet_name='Consolidado_Criticas', index=False, startrow=2)
            
            # ABA 3: Ranking por Polo
            ranking = self._gerar_ranking_polos(relatorios_polo)
            ranking.to_excel(writer, sheet_name='Ranking_Polos', index=False, startrow=2)
            
            # Aplicar formatação
            self._aplicar_formatacao_consolidado(writer)
        
        return output.getvalue()
    
    def _gerar_resumo_polo(self, dados: pd.DataFrame, nome_polo: str) -> pd.DataFrame:
        """Gera resumo executivo detalhado do polo"""
        
        resumo = []
        
        # Cabeçalho
        resumo.extend([
            {'Categoria': 'RELATÓRIO EXECUTIVO - POLO', 'Valor': nome_polo},
            {'Categoria': 'Data/Hora da Geração', 'Valor': datetime.now().strftime('%d/%m/%Y %H:%M')},
            {'Categoria': 'Total de Registros', 'Valor': len(dados)},
            {'Categoria': '', 'Valor': ''},
        ])
        
        # Distribuição por urgência
        if 'Nivel_Urgencia' in dados.columns:
            resumo.append({'Categoria': 'DISTRIBUIÇÃO POR URGÊNCIA', 'Valor': ''})
            dist_urgencia = dados['Nivel_Urgencia'].value_counts().sort_index(ascending=False)
            total_ordens = len(dados)
            
            for nivel, qtd in dist_urgencia.items():
                descricao = {5: '🔴 CRÍTICO', 4: '🟠 ALTO', 3: '🟡 MÉDIO', 2: '🔵 BAIXO', 1: '⚪ NORMAL'}.get(nivel, 'N/A')
                percentual = (qtd/total_ordens*100) if total_ordens > 0 else 0
                resumo.append({'Categoria': f'  {descricao}', 'Valor': f"{qtd} ({percentual:.1f}%)"})
            resumo.append({'Categoria': '', 'Valor': ''})
        
        # Distribuição por Status SLA
        if 'Status_SLA' in dados.columns:
            resumo.append({'Categoria': 'DISTRIBUIÇÃO POR STATUS SLA', 'Valor': ''})
            dist_sla = dados['Status_SLA'].value_counts()
            total_ordens = len(dados)
            
            for status, qtd in dist_sla.items():
                percentual = (qtd/total_ordens*100) if total_ordens > 0 else 0
                resumo.append({'Categoria': f'  {status}', 'Valor': f"{qtd} ({percentual:.1f}%)"})
            resumo.append({'Categoria': '', 'Valor': ''})
        
        # Estatísticas de dias em aberto
        if 'Dias_Em_Aberto' in dados.columns:
            resumo.extend([
                {'Categoria': 'ESTATÍSTICAS DIAS EM ABERTO', 'Valor': ''},
                {'Categoria': '  Média', 'Valor': f"{dados['Dias_Em_Aberto'].mean():.1f} dias"},
                {'Categoria': '  Mediana', 'Valor': f"{dados['Dias_Em_Aberto'].median():.1f} dias"},
                {'Categoria': '  Máximo', 'Valor': f"{dados['Dias_Em_Aberto'].max()} dias"},
                {'Categoria': '  Mínimo', 'Valor': f"{dados['Dias_Em_Aberto'].min()} dias"},
                {'Categoria': '  Desvio Padrão', 'Valor': f"{dados['Dias_Em_Aberto'].std():.1f} dias"},
                {'Categoria': '', 'Valor': ''},
            ])
        
        # Distribuição por Estado (Top 10)
        if 'Estado' in dados.columns:
            resumo.append({'Categoria': 'TOP 10 ESTADOS', 'Valor': ''})
            dist_estados = dados['Estado'].value_counts().head(10)
            total_ordens = len(dados)
            
            for estado, qtd in dist_estados.items():
                percentual = (qtd/total_ordens*100) if total_ordens > 0 else 0
                resumo.append({'Categoria': f'  {estado}', 'Valor': f"{qtd} ({percentual:.1f}%)"})
        
        return pd.DataFrame(resumo)
    
    def _gerar_estatisticas_detalhadas(self, dados: pd.DataFrame) -> pd.DataFrame:
        """Gera estatísticas detalhadas para análise"""
        
        stats = []
        
        # Estatísticas básicas
        stats.extend([
            {'Categoria': 'ESTATÍSTICAS GERAIS', 'Métrica': 'Total de Registros', 'Valor': len(dados)},
            {'Categoria': 'ESTATÍSTICAS GERAIS', 'Métrica': 'Data da Análise', 'Valor': datetime.now().strftime('%d/%m/%Y %H:%M')},
        ])
        
        # Estatísticas de Dias em Aberto
        if 'Dias_Em_Aberto' in dados.columns:
            dias_stats = dados['Dias_Em_Aberto'].describe()
            for stat, valor in dias_stats.items():
                nome_stat = {
                    'count': 'Quantidade de Registros',
                    'mean': 'Média',
                    'std': 'Desvio Padrão',
                    'min': 'Mínimo',
                    '25%': 'Percentil 25%',
                    '50%': 'Mediana (Percentil 50%)',
                    '75%': 'Percentil 75%',
                    'max': 'Máximo'
                }.get(stat, stat)
                
                stats.append({'Categoria': 'DIAS EM ABERTO', 'Métrica': nome_stat, 'Valor': f"{valor:.2f}"})
        
        # Estatísticas de SLA Cliente
        if 'SLA Cliente' in dados.columns:
            sla_stats = dados['SLA Cliente'].describe()
            for stat, valor in sla_stats.items():
                nome_stat = {
                    'count': 'Quantidade de Registros',
                    'mean': 'SLA Médio',
                    'std': 'Desvio Padrão SLA',
                    'min': 'SLA Mínimo',
                    '25%': 'SLA Percentil 25%',
                    '50%': 'SLA Mediano',
                    '75%': 'SLA Percentil 75%',
                    'max': 'SLA Máximo'
                }.get(stat, stat)
                
                stats.append({'Categoria': 'SLA CLIENTE', 'Métrica': nome_stat, 'Valor': f"{valor:.2f}"})
        
        # Análise de correlação (se possível)
        if 'Dias_Em_Aberto' in dados.columns and 'SLA Cliente' in dados.columns:
            try:
                correlacao = dados['Dias_Em_Aberto'].corr(dados['SLA Cliente'])
                stats.append({'Categoria': 'CORRELAÇÃO', 'Métrica': 'Dias em Aberto vs SLA Cliente', 'Valor': f"{correlacao:.3f}"})
            except:
                pass
        
        return pd.DataFrame(stats)
    
    def _gerar_analise_geografica(self, dados: pd.DataFrame) -> pd.DataFrame:
        """Gera análise geográfica detalhada"""
        
        analise = []
        
        if 'Estado' in dados.columns:
            # Análise por Estado
            analise.append({'Tipo': 'DISTRIBUIÇÃO POR ESTADO', 'Local': '', 'Quantidade': '', 'Percentual': '', 'Média_Dias': ''})
            
            dist_estados = dados.groupby('Estado').agg({
                'Ordem PagBank': 'count',
                'Dias_Em_Aberto': 'mean'
            }).round(1)
            
            dist_estados.columns = ['Quantidade', 'Média_Dias']
            dist_estados['Percentual'] = (dist_estados['Quantidade'] / len(dados) * 100).round(1)
            dist_estados = dist_estados.sort_values('Quantidade', ascending=False)
            
            for estado, row in dist_estados.iterrows():
                analise.append({
                    'Tipo': 'Estado',
                    'Local': estado,
                    'Quantidade': int(row['Quantidade']),
                    'Percentual': f"{row['Percentual']:.1f}%",
                    'Média_Dias': f"{row['Média_Dias']:.1f}"
                })
        
        if 'Cidade' in dados.columns:
            # Top 20 Cidades
            analise.append({'Tipo': '', 'Local': '', 'Quantidade': '', 'Percentual': '', 'Média_Dias': ''})
            analise.append({'Tipo': 'TOP 20 CIDADES', 'Local': '', 'Quantidade': '', 'Percentual': '', 'Média_Dias': ''})
            
            dist_cidades = dados.groupby('Cidade').agg({
                'Ordem PagBank': 'count',
                'Dias_Em_Aberto': 'mean'
            }).round(1)
            
            dist_cidades.columns = ['Quantidade', 'Média_Dias']
            dist_cidades['Percentual'] = (dist_cidades['Quantidade'] / len(dados) * 100).round(1)
            dist_cidades = dist_cidades.sort_values('Quantidade', ascending=False).head(20)
            
            for cidade, row in dist_cidades.iterrows():
                analise.append({
                    'Tipo': 'Cidade',
                    'Local': cidade,
                    'Quantidade': int(row['Quantidade']),
                    'Percentual': f"{row['Percentual']:.1f}%",
                    'Média_Dias': f"{row['Média_Dias']:.1f}"
                })
        
        return pd.DataFrame(analise)
    
    def _gerar_resumo_geral_polos(self, relatorios_polo: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """Gera resumo geral de todos os polos"""
        
        resumo_geral = []
        
        for polo_id, dados_polo in relatorios_polo.items():
            total = len(dados_polo)
            criticas = len(dados_polo[dados_polo['Nivel_Urgencia'] == 5]) if 'Nivel_Urgencia' in dados_polo.columns else 0
            altas = len(dados_polo[dados_polo['Nivel_Urgencia'] == 4]) if 'Nivel_Urgencia' in dados_polo.columns else 0
            vencidas = len(dados_polo[dados_polo['Status_SLA'] == 'Vencido']) if 'Status_SLA' in dados_polo.columns else 0
            media_dias = dados_polo['Dias_Em_Aberto'].mean() if 'Dias_Em_Aberto' in dados_polo.columns else 0
            max_dias = dados_polo['Dias_Em_Aberto'].max() if 'Dias_Em_Aberto' in dados_polo.columns else 0
            
            resumo_geral.append({
                'Polo': polo_id,
                'Total_Ordens': total,
                'Ordens_Criticas': criticas,
                'Ordens_Altas': altas,
                'SLA_Vencido': vencidas,
                'Perc_Criticas': f"{(criticas/total*100):.1f}%" if total > 0 else "0%",
                'Perc_Vencidas': f"{(vencidas/total*100):.1f}%" if total > 0 else "0%",
                'Media_Dias_Aberto': f"{media_dias:.1f}",
                'Max_Dias_Aberto': max_dias,
                'Data_Analise': datetime.now().strftime('%d/%m/%Y %H:%M')
            })
        
        return pd.DataFrame(resumo_geral)
    
    def _gerar_ranking_polos(self, relatorios_polo: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """Gera ranking de performance dos polos"""
        
        ranking = []
        
        for polo_id, dados_polo in relatorios_polo.items():
            total = len(dados_polo)
            criticas = len(dados_polo[dados_polo['Nivel_Urgencia'] == 5]) if 'Nivel_Urgencia' in dados_polo.columns else 0
            vencidas = len(dados_polo[dados_polo['Status_SLA'] == 'Vencido']) if 'Status_SLA' in dados_polo.columns else 0
            media_dias = dados_polo['Dias_Em_Aberto'].mean() if 'Dias_Em_Aberto' in dados_polo.columns else 0
            
            # Calcular score de performance (menor é melhor)
            score_criticas = (criticas / total * 100) if total > 0 else 0
            score_vencidas = (vencidas / total * 100) if total > 0 else 0
            score_media_dias = media_dias
            
            # Score final (ponderado)
            score_final = (score_criticas * 0.4) + (score_vencidas * 0.4) + (score_media_dias * 0.2)
            
            ranking.append({
                'Polo': polo_id,
                'Score_Performance': round(score_final, 2),
                'Perc_Criticas': f"{score_criticas:.1f}%",
                'Perc_Vencidas': f"{score_vencidas:.1f}%",
                'Media_Dias': f"{media_dias:.1f}",
                'Total_Ordens': total,
                'Status_Performance': self._classificar_performance(score_final)
            })
        
        # Ordenar por score (menor é melhor)
        df_ranking = pd.DataFrame(ranking).sort_values('Score_Performance')
        df_ranking['Posicao'] = range(1, len(df_ranking) + 1)
        
        # Reordenar colunas
        df_ranking = df_ranking[['Posicao', 'Polo', 'Status_Performance', 'Score_Performance', 
                                'Total_Ordens', 'Perc_Criticas', 'Perc_Vencidas', 'Media_Dias']]
        
        return df_ranking
    
    def _classificar_performance(self, score: float) -> str:
        """Classifica performance baseada no score"""
        if score <= 10:
            return '🟢 EXCELENTE'
        elif score <= 25:
            return '🟡 BOM'
        elif score <= 50:
            return '🟠 REGULAR'
        else:
            return '🔴 CRÍTICO'
    
    def _aplicar_formatacao_excel_avancada(self, writer, nome_polo: str, dados_polo: pd.DataFrame):
        """Aplica formatação avançada ao Excel"""
        try:
            workbook = writer.book
            
            # Formatação para cada aba
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Título principal
                worksheet.merge_cells('A1:J1')
                title_cell = worksheet['A1']
                title_cell.value = f"RELATÓRIO {sheet_name.upper().replace('_', ' ')} - {nome_polo}"
                title_cell.font = Font(size=14, bold=True, color='FFFFFF')
                title_cell.fill = PatternFill(start_color=self.cores['header'], end_color=self.cores['header'], fill_type='solid')
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Autoajustar largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Formatação do cabeçalho (linha 3)
                if worksheet.max_row >= 3:
                    for cell in worksheet[3]:
                        if cell.value:
                            cell.font = Font(bold=True, color='FFFFFF')
                            cell.fill = PatternFill(start_color=self.cores['subheader'], end_color=self.cores['subheader'], fill_type='solid')
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Formatação condicional para urgência (se aplicável)
                if sheet_name in ['Ordens_Detalhadas', 'Ordens_Criticas']:
                    self._aplicar_formatacao_condicional_urgencia(worksheet, dados_polo)
                
                # Bordas para todas as células com dados
                if worksheet.max_row > 3:
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, 
                                                 min_col=1, max_col=worksheet.max_column):
                        for cell in row:
                            if cell.value is not None:
                                cell.border = thin_border
        
        except Exception as e:
            # Se falhar formatação, continua sem ela
            print(f"Aviso: Erro na formatação Excel: {e}")
    
    def _aplicar_formatacao_condicional_urgencia(self, worksheet, dados_polo: pd.DataFrame):
        """Aplica formatação condicional baseada na urgência"""
        try:
            # Encontrar coluna de urgência
            urgencia_col = None
            for idx, cell in enumerate(worksheet[3], 1):
                if cell.value and 'urgencia' in str(cell.value).lower():
                    urgencia_col = idx
                    break
            
            if urgencia_col:
                # Aplicar cores baseadas na urgência
                for row_idx in range(4, worksheet.max_row + 1):
                    urgencia_cell = worksheet.cell(row=row_idx, column=urgencia_col)
                    if urgencia_cell.value:
                        urgencia_text = str(urgencia_cell.value).lower()
                        
                        if 'crítico' in urgencia_text:
                            fill_color = self.cores['critico']
                        elif 'alto' in urgencia_text:
                            fill_color = self.cores['alto']
                        elif 'médio' in urgencia_text:
                            fill_color = self.cores['medio']
                        elif 'baixo' in urgencia_text:
                            fill_color = self.cores['baixo']
                        else:
                            fill_color = self.cores['normal']
                        
                        # Aplicar cor à linha inteira
                        for col_idx in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        except Exception:
            pass  # Ignora erros de formatação condicional
    
    def _aplicar_formatacao_consolidado(self, writer):
        """Aplica formatação específica para relatório consolidado"""
        try:
            workbook = writer.book
            
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Título principal
                worksheet.merge_cells('A1:H1')
                title_cell = worksheet['A1']
                title_cell.value = f"RELATÓRIO CONSOLIDADO - {sheet_name.upper().replace('_', ' ')}"
                title_cell.font = Font(size=16, bold=True, color='FFFFFF')
                title_cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Autoajustar colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 40)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        except Exception:
            pass  # Ignora erros de formatação
